# import os
#
# print(os.listdir("../ch21"))

# from pathlib import Path
#
# folder_path = Path("./test")
#
# txt_files = folder_path.glob("*.txt")
#
# for txt_file in txt_files:
#     print(txt_file.name)
#
# new_folder = Path("new_folder")
#
# new_folder.mkdir(exist_ok=True)


from openpyxl import Workbook, load_workbook

# wb = Workbook()
#
# wb.save("data.xlsx")

wb = load_workbook("data.xlsx")

ws = wb.worksheets[0]

ws.append(["이름", "나이", "성별"])
ws.append(["철수", 10, "남자"])
ws.append(["영희", 12, "여자"])

# 저장
wb.save("data.xlsx")
